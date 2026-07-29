"""GET /api/reports/tickets/export — exportación a Excel (spec 034, US5).

Ultra-limitado (Principio VII): exactamente 5 tickets mock, acotados por client_id/project_id.
"""
import io

from openpyxl import load_workbook


def test_exporta_columnas_orden_y_filas_del_conjunto_filtrado(client, ticket_client, ticket_project, make_ticket):
    for i in range(5):
        make_ticket(title=f"Ticket export {i}")

    response = client.get("/api/reports/tickets/export", query_string=[
        ("client_id", ticket_client["id"]),
        ("project_id", ticket_project["id"]),
        ("columns", "title"),
        ("columns", "status"),
    ])

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(io.BytesIO(response.data))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert header == ["Título", "Estado"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 5
    assert {row[1] for row in data_rows} == {"nuevo"}


def test_exportar_sin_filas_devuelve_no_data(client, ticket_client):
    import uuid
    response = client.get("/api/reports/tickets/export", query_string={
        "client_id": ticket_client["id"], "assignee_id": str(uuid.uuid4()),
    })

    assert response.status_code == 400
    assert response.get_json()["error"] == "no_data"


def test_exportar_columna_desconocida_es_rechazada(client, ticket_client):
    response = client.get("/api/reports/tickets/export", query_string={
        "client_id": ticket_client["id"], "columns": "no_existe",
    })

    assert response.status_code == 400
    assert response.get_json()["error"] == "validation_error"
